import os
import io
import uuid
import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from supabase import create_client, Client
import json
from datetime import datetime, timezone
from dotenv import load_dotenv

load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_SUPER_KEY")
SERVICE_ACCOUNT_PATH = os.getenv("GCP_SERVICE_ACCOUNT_PATH")

if not SERVICE_ACCOUNT_PATH or not os.path.exists(SERVICE_ACCOUNT_PATH):
    raise FileNotFoundError(f"Service account file not found at: {SERVICE_ACCOUNT_PATH}")

SCOPES = ['https://www.googleapis.com/auth/drive.readonly']
creds = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_PATH, scopes=SCOPES)
drive = build("drive", "v3", credentials=creds)

try:
    with open(SERVICE_ACCOUNT_PATH, 'r') as f:
        SERVICE_ACCOUNT_INFO = json.load(f)
except Exception as e:
    raise RuntimeError(f"Failed to load service account info from {SERVICE_ACCOUNT_PATH}: {e}")


def enrich_assignment_names(assignments: list, rad_map: dict, fac_map: dict) -> list:
    enriched = []
    for a in assignments:
        enriched.append({
            **a,
            "radiologist_name": rad_map.get(a["radiologist_id"]),
            "facility_name": fac_map.get(a.get("facility_id"), "N/A")
        })
    return enriched


def process_cell_update(sheet_id: str, row: int, col: int) -> tuple[dict, int]:
    print(f"Starting radmapping sync for sheet={sheet_id}, row={row}, col={col}")

    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
    _creds_local = service_account.Credentials.from_service_account_info(SERVICE_ACCOUNT_INFO, scopes=SCOPES)
    _drive_local = build("drive", "v3", credentials=_creds_local)

    temp_file = f"/tmp/radmapping_{uuid.uuid4()}.xlsx"

    print(f"Downloading Google Sheet to {temp_file}...")
    request = _drive_local.files().export_media( 
        fileId=sheet_id,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    try:
        with open(temp_file, "wb") as f:
            downloader = MediaIoBaseDownload(f, request)
            done = False
            while not done:
                status, done = downloader.next_chunk()
                print(f"Download progress: {int(status.progress() * 100)}%")
    except Exception as e:
        print(f"Error downloading file: {e}")
        return {"error": f"Failed to download Google Sheet: {e}"}, 500


    try:
        wb = openpyxl.load_workbook(temp_file)
        ws = wb["RAD MAPPING "] 
    except KeyError:
        return {"error": "Sheet 'RAD MAPPING ' not found in the workbook."}, 400
    except Exception as e:
        return {"error": f"Error loading workbook: {e}"}, 500
    finally:
        if os.path.exists(temp_file):
            os.remove(temp_file)
            print(f"Temp file deleted.")


    facility_name = str(ws.cell(row=row, column=1).value).strip().upper()
    if not facility_name:
        return {"error": "Missing facility name in column A"}, 400

    facilities_res = supabase.table("facilities").select("id, name").execute()
    if facilities_res.data is None:
        return {"error": "Failed to fetch facilities from Supabase."}, 500
    facilities = facilities_res.data
    facility_lookup = {f["name"].strip().upper(): f["id"] for f in facilities}
    facility_id = facility_lookup.get(facility_name)
    if not facility_id:
        return {"error": f"Facility not found: {facility_name}"}, 404

    print(f"Matched facility: {facility_name} → {facility_id}")

    radiologists_res = supabase.table("radiologists").select("id, name").execute()
    if radiologists_res.data is None:
        return {"error": "Failed to fetch radiologists from Supabase."}, 500
    radiologists = radiologists_res.data
    rad_lookup = {}
    for rad in radiologists:
        for part in rad["name"].upper().replace(",", "").split():
            rad_lookup.setdefault(part, []).append(rad["id"])

    seen_rads_in_sheet = set()
    new_assignments_from_sheet = []
    
    COLOR_RED = "FFFF0000"
    COLOR_BLUE = "FF0000FF"
    COLOR_MAGENTA = "FFFF00FF"

    for col_idx in range(2, ws.max_column + 1):
        cell = ws.cell(row=row, column=col_idx)
        value = str(cell.value).strip() if cell.value else ""
        if not value:
            continue

        parts = value.split()
        name_part = parts[0].strip(",").upper()
        notes = " ".join(parts[1:]) if len(parts) > 1 else ""
        matching = rad_lookup.get(name_part, [])

        if len(matching) != 1:
            print(f"Could not uniquely match doctor: {name_part} (Matches found: {len(matching)})")
            continue

        rad_id = matching[0]
        seen_rads_in_sheet.add(rad_id)

        font_color = cell.font.color
        if font_color and font_color.type == "rgb":
            rgb = font_color.rgb.upper()
            if rgb.startswith(COLOR_RED):
                can_read = "false"
            elif rgb.startswith(COLOR_BLUE):
                can_read = "pending"
            elif rgb.startswith(COLOR_MAGENTA):
                can_read = "withdrawn"
            else:
                can_read = "true"
        else:
            can_read = "true"

        new_assignments_from_sheet.append({
            "radiologist_id": rad_id,
            "facility_id": facility_id,
            "can_read": can_read,
            "notes": notes
        })

    existing_assignments_res = supabase.table("doctor_facility_assignments") \
        .select("id, radiologist_id, facility_id, can_read, notes") \
        .eq("facility_id", facility_id).execute()

    if existing_assignments_res.data is None:
        return {"error": "Failed to fetch existing assignments from Supabase."}, 500
    
    existing_assignments = existing_assignments_res.data
    existing_assignments_map = {a["radiologist_id"]: a for a in existing_assignments}

    assignments_added = []
    assignments_removed = []
    assignments_updated = [] 
    for rad_id, assignment in existing_assignments_map.items():
        if rad_id not in seen_rads_in_sheet:
            print(f"Removing stale assignment for rad_id={rad_id}, facility_id={facility_id}")
            supabase.table("doctor_facility_assignments") \
                .delete().eq("radiologist_id", rad_id).eq("facility_id", facility_id).execute()
            assignments_removed.append(assignment)

    for new_assignment in new_assignments_from_sheet:
        rad_id = new_assignment["radiologist_id"]
        existing = existing_assignments_map.get(rad_id)

        if existing:
            changed = (
                new_assignment["can_read"] != existing["can_read"] or
                new_assignment["notes"] != (existing["notes"] or "")
            )
            if changed:
                print(f"Updating assignment for rad_id={rad_id}, facility_id={facility_id}")
                supabase.table("doctor_facility_assignments") \
                    .update(new_assignment) \
                    .eq("radiologist_id", rad_id) \
                    .eq("facility_id", facility_id) \
                    .execute()
                assignments_removed.append(existing) 
                assignments_added.append({**new_assignment, "id": existing["id"]}) 
                assignments_updated.append(new_assignment) 
            else:
                print(f"Skipping unchanged assignment for rad_id={rad_id}, facility_id={facility_id}")
        else:
            print(f"Adding new assignment for rad_id={rad_id}, facility_id={facility_id}")
            new_record_id = str(uuid.uuid4())
            record_to_insert = {
                "id": new_record_id,
                **new_assignment
            }
            supabase.table("doctor_facility_assignments").insert(record_to_insert).execute()
            assignments_added.append(record_to_insert)

    rad_name_map = {r["id"]: r["name"] for r in radiologists}
    fac_name_map = {f["id"]: f["name"] for f in facilities}

    assignments_added_named = enrich_assignment_names(assignments_added, rad_name_map, fac_name_map)
    assignments_removed_named = enrich_assignment_names(assignments_removed, rad_name_map, fac_name_map)

    audit_entry = {
        "action": "sync",
        "table_name": "doctor_facility_assignments",
        "record_id": facility_id, 
        "user_email": "radmapping@sync",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "old_data": {"assignments_removed": assignments_removed_named},
        "new_data": {
            "assignments_added": assignments_added_named,
            "summary": {
                "added": len(assignments_added), 
                "removed": len(assignments_removed),
                "updated": len(assignments_updated) 
            }
        }
    }

    try:
        supabase.table("audit_log").insert(audit_entry).execute()
        print("Audit log entry created successfully.")
    except Exception as e:
        print(f"Failed to create audit log entry: {e}")

    return {
        "status": "success",
        "facility": facility_name,
        "assignments_added": len(assignments_added),
        "assignments_removed": len(assignments_removed),
        "assignments_updated": len(assignments_updated)
    }, 200